from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from io import BytesIO


def create_export_file(export_dict: dict):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "My_sheet1"
    sheet2 = wb.create_sheet(title="My_sheet2")
    i = 1
    for k, v in export_dict.items():
        sheet2.cell(row=i, column=1).value = k
        sheet2.cell(row=i, column=2).value = v
        i+=1

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = BytesIO(tmp.read())
    return output
